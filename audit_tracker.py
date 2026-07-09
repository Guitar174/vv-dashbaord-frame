"""
audit_tracker.py — Data reconciliation audit.
Auto-rebuilds if source is newer than dashboard, then audits.
Writes findings to audit_log.txt.
"""
import openpyxl, json, re, datetime, subprocess, sys
from pathlib import Path
from collections import Counter

SRC_DIR   = Path(__file__).parent / "Source"
HTML_PATH = Path(__file__).parent / "sales_order_tracker.html"
LOG_PATH  = Path(__file__).parent / "audit_log.txt"
TODAY     = datetime.date.today()

STATUS_ORDER = ['Open', 'Pending Approval', 'Released', 'Closed', 'Cancelled']

lines = []
def log(s=""): lines.append(s); print(s)

log("=" * 60)
log(f"DATA AUDIT — {TODAY}")
log("=" * 60)

# ── Load source Excel ────────────────────────────────────────────
files = sorted(
    [f for f in SRC_DIR.glob("*.xlsx") if f.name != "Sale list.xlsx"],
    key=lambda f: f.stat().st_mtime
)
xl = files[-1]
# Check if source file name is not already embedded in dashboard
html_check = HTML_PATH.read_text(encoding="utf-8")
if xl.name not in html_check:
    log(f"Source file not in dashboard — rebuilding...")
    subprocess.run([sys.executable, Path(__file__).parent / "rebuild_tracker.py"], check=True)
log(f"Source file : {xl.name}")

wb = openpyxl.load_workbook(xl, data_only=True)
ws = wb["Sales Orders"]

nick_map = {}
sale_list = SRC_DIR / "Sale list.xlsx"
if sale_list.exists():
    wb2 = openpyxl.load_workbook(sale_list, data_only=True)
    for row in wb2["Frame"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            nick_map[str(row[0]).strip()] = str(row[1]).strip()

header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
def col(name): return header.index(name)
C_NO, C_SALES, C_AMOUNT = col("No."), col("Salesperson Code"), col("Amount Including VAT")
C_DATE, C_STATUS, C_SCODE = col("Document Date"), col("Status"), col("Status Code")

src_rows = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[C_NO]:
        continue
    oid       = str(row[C_NO]).strip()
    raw_code  = str(row[C_SALES] or "").strip()
    nickname  = nick_map.get(raw_code, raw_code)
    amt       = float(row[C_AMOUNT]) if row[C_AMOUNT] is not None else 0.0
    doc_date  = row[C_DATE]
    status    = str(row[C_STATUS] or "").strip()
    s_code    = str(row[C_SCODE] or "").strip()
    if isinstance(doc_date, (datetime.datetime, datetime.date)):
        date_str = doc_date.strftime("%d/%m/%Y")
    else:
        date_str = str(doc_date or "")
    src_rows[oid] = {"sales": nickname, "sales_code": raw_code,
                     "amount": amt, "date": date_str,
                     "status": status, "issue": s_code}

# ── Load dashboard ORDERS ────────────────────────────────────────
html = HTML_PATH.read_text(encoding="utf-8")
m = re.search(r"const ORDERS\s*=\s*(\[[\s\S]*?\]);", html)
dash_rows = {o["id"]: o for o in json.loads(m.group(1))}

log(f"Source rows : {len(src_rows)}")
log(f"Dashboard   : {len(dash_rows)}")
log()

findings = []

# CHECK 1: Row count & missing IDs
log("CHECK 1 — ROW COUNT")
only_src  = [i for i in src_rows  if i not in dash_rows]
only_dash = [i for i in dash_rows if i not in src_rows]
if not only_src and not only_dash and len(src_rows) == len(dash_rows):
    log("  PASS — counts match, no missing IDs")
else:
    msg = f"FAIL — src={len(src_rows)} dash={len(dash_rows)}"
    if only_src:  msg += f" | in src only: {only_src[:10]}"
    if only_dash: msg += f" | in dash only: {only_dash[:10]}"
    log(f"  {msg}")
    findings.append(("ROW COUNT", msg))
log()

# CHECK 2: Status breakdown
log("CHECK 2 — STATUS BREAKDOWN")
src_status  = Counter(v["status"] for v in src_rows.values())
dash_status = Counter(v["status"] for v in dash_rows.values())
all_statuses = set(src_status) | set(dash_status)
fail2 = []
for s in sorted(all_statuses):
    sv, dv = src_status.get(s, 0), dash_status.get(s, 0)
    ok = "OK" if sv == dv else "!!"
    log(f"  {ok} {s:20s}  src={sv}  dash={dv}")
    if sv != dv:
        fail2.append(f"{s}: src={sv} dash={dv}")
if fail2:
    findings.append(("STATUS BREAKDOWN", "; ".join(fail2)))
log()

# CHECK 3: Amount totals
log("CHECK 3 — AMOUNT TOTALS")
src_total  = sum(v["amount"] for v in src_rows.values())
dash_total = sum(v["amount"] for v in dash_rows.values())
diff = abs(src_total - dash_total)
if diff < 0.01:
    log(f"  PASS — total ฿{src_total:,.2f}")
else:
    msg = f"FAIL — src=฿{src_total:,.2f} dash=฿{dash_total:,.2f} diff=฿{diff:,.2f}"
    log(f"  {msg}")
    findings.append(("AMOUNT TOTAL", msg))

# Sample 20 rows
common = list(set(src_rows) & set(dash_rows))[:20]
amt_mismatches = []
for oid in common:
    sa, da = src_rows[oid]["amount"], float(dash_rows[oid]["amount"])
    if abs(sa - da) > 0.01:
        amt_mismatches.append(f"{oid}: src={sa} dash={da}")
if amt_mismatches:
    log(f"  Sample mismatches: {amt_mismatches}")
    findings.append(("AMOUNT SAMPLE", "; ".join(amt_mismatches)))
else:
    log(f"  Sample (20 rows): all match")
log()

# CHECK 4: Date field (10 rows)
log("CHECK 4 — DATE FIELD (10 rows)")
date_fails = []
for oid in list(common)[:10]:
    sd, dd = src_rows[oid]["date"], dash_rows[oid].get("date", "")
    if sd != dd:
        date_fails.append(f"{oid}: src={sd} dash={dd}")
if date_fails:
    log(f"  FAIL — {date_fails}")
    findings.append(("DATE FIELD", "; ".join(date_fails)))
else:
    log("  PASS")
log()

# CHECK 5: Duplicate IDs in dashboard
log("CHECK 5 — DUPLICATE ORDER IDs")
all_ids = [o["id"] for o in json.loads(m.group(1))]
dup_counts = {i: c for i, c in Counter(all_ids).items() if c > 1}
if dup_counts:
    log(f"  FAIL — {dup_counts}")
    findings.append(("DUPLICATES", str(dup_counts)))
else:
    log("  PASS — no duplicates")
log()

# CHECK 6: Sales code → nickname mapping
log("CHECK 6 — SALES CODE MAPPING")
map_fails = []
for oid in list(common)[:50]:
    expected_nick = nick_map.get(src_rows[oid]["sales_code"], src_rows[oid]["sales_code"])
    dash_nick = dash_rows[oid].get("sales", "")
    if expected_nick != dash_nick:
        map_fails.append(f"{oid}: expected={expected_nick} dash={dash_nick}")
if map_fails:
    log(f"  FAIL — {map_fails[:5]}")
    findings.append(("SALES MAPPING", "; ".join(map_fails[:5])))
else:
    log("  PASS")
log()

# CHECK 7: Cancelled with non-empty status code
log("CHECK 7 — CANCELLED WITH ACTIVE STATUS CODE")
contradictions = [
    oid for oid, v in dash_rows.items()
    if v.get("status") == "Cancelled" and v.get("issue", "").strip()
]
if contradictions:
    log(f"  NOTE — {len(contradictions)} Cancelled rows have a status code (may be normal)")
    log(f"  Sample: {contradictions[:5]}")
else:
    log("  PASS — no contradictions")
log()

# ── Summary ──────────────────────────────────────────────────────
log("=" * 60)
log("TOP FINDINGS (by dashboard impact)")
log("=" * 60)
if findings:
    for i, (title, detail) in enumerate(findings[:3], 1):
        log(f"{i}. [{title}] {detail}")
else:
    log("All checks passed — dashboard matches source.")
log()

LOG_PATH.write_text("\n".join(lines), encoding="utf-8")
print(f"\nAudit saved to: {LOG_PATH.name}")
