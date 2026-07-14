"""
rebuild_tracker.py  —  อัพเดทข้อมูลใน sales_order_tracker.html
จากไฟล์ Excel ล่าสุดใน Source/
"""
import openpyxl, json, re, datetime
from pathlib import Path

SRC_DIR  = Path(__file__).parent / "Source"
HTML_PATH = Path(__file__).parent / "sales_order_tracker.html"

# ── 1. เลือกไฟล์ล่าสุด (ยกเว้น Sale list.xlsx) ──────────────────────────
files = sorted(
    [f for f in SRC_DIR.glob("*.xlsx") if f.name != "Sale list.xlsx"],
    key=lambda f: f.stat().st_mtime
)
if not files:
    raise FileNotFoundError("ไม่พบไฟล์ .xlsx ใน Source/")

# หาไฟล์ใหม่สุดที่มี sheet "Sales Orders" (บางไฟล์ export มาไม่ครบ)
xl = None
for f in reversed(files):
    try:
        _wb = openpyxl.load_workbook(f, read_only=True)
        has_sheet = "Sales Orders" in _wb.sheetnames
        _wb.close()
        if has_sheet:
            xl = f
            break
        print(f"Skip {f.name}: no 'Sales Orders' sheet")
    except Exception as e:
        print(f"Skip {f.name}: {e}")
if xl is None:
    raise FileNotFoundError("ไม่พบไฟล์ที่มี sheet 'Sales Orders' ใน Source/")
print(f"Using: {xl.name}")

# ── 2. โหลด nickname map จาก Sale list.xlsx ─────────────────────────────
nick_map = {}
sale_list = SRC_DIR / "Sale list.xlsx"
if sale_list.exists():
    wb2 = openpyxl.load_workbook(sale_list, data_only=True)
    for row in wb2["Frame"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            nick_map[str(row[0]).strip()] = str(row[1]).strip()
nick_to_code = {v: k for k, v in nick_map.items()}

# ── 3. อ่านข้อมูล (หา column จากชื่อ header — กัน layout เปลี่ยน) ────────
wb = openpyxl.load_workbook(xl, data_only=True)
ws = wb["Sales Orders"]

header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
def col(name):
    if name not in header:
        raise ValueError(f"Column '{name}' not found in {xl.name}. Headers: {header}")
    return header.index(name)

C_NO, C_STATUS_CODE, C_STATUS = col("No."), col("Status Code"), col("Status")
C_CUST, C_AMOUNT, C_DATE      = col("Sell-to Customer Name"), col("Amount Including VAT"), col("Document Date")
C_SALES, C_COMMENT, C_WORK    = col("Salesperson Code"), col("First Comment"), col("Work Description")

orders    = []
sales_set = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    inv_no = row[C_NO]
    if not inv_no:
        continue

    customer      = str(row[C_CUST] or "").strip()
    raw_code      = str(row[C_SALES] or "").strip()
    sales_code    = nick_map.get(raw_code, raw_code)       # map to nickname
    amount        = float(row[C_AMOUNT]) if row[C_AMOUNT] is not None else 0.0
    doc_date      = row[C_DATE]
    status        = str(row[C_STATUS] or "").strip()
    status_code   = str(row[C_STATUS_CODE] or "").strip()
    first_comment = re.sub(r"[\r\n\t]+", " ", str(row[C_COMMENT] or "")).strip()
    work_desc     = re.sub(r"[\r\n\t]+", " ", str(row[C_WORK] or "")).strip()

    # Format date DD/MM/YYYY
    if isinstance(doc_date, (datetime.datetime, datetime.date)):
        date_str = doc_date.strftime("%d/%m/%Y")
    else:
        date_str = str(doc_date or "")

    sales_set.add(sales_code)
    orders.append({
        "id":            str(inv_no).strip(),
        "customer":      customer,
        "sales":         sales_code,
        "sales_code":    raw_code,
        "amount":        amount,
        "date":          date_str,
        "status":        status,
        "issue":         status_code,
        "first_comment": first_comment,
        "work_desc":     work_desc,
    })

print(f"Loaded {len(orders)} rows | {len(sales_set)} salespeople: {sorted(sales_set)}")

# ── 3. Inject ข้อมูลลง HTML ─────────────────────────────────────────────
html = HTML_PATH.read_text(encoding="utf-8")

orders_js = "const ORDERS = " + json.dumps(orders, ensure_ascii=False, separators=(",", ":")) + ";"

# Replace ORDERS block
html = re.sub(r"const ORDERS\s*=\s*\[[\s\S]*?\];", orders_js, html)

# ── 4. อัพเดท salesperson options (เฉพาะ id="ms-sales-options") ─────────
sales_options_html = "\n".join(
    f'          <label class="ms-option"><span>{nick_to_code.get(code, "")} ({code})</span>'
    f'<input type="checkbox" value="{code}" onchange="onMsChange(\'ms-sales\')"></label>'
    for code in sorted(sales_set, key=lambda n: nick_to_code.get(n, n))
)
html = re.sub(
    r'(<div class="ms-options"[^>]*id="ms-sales-options"[^>]*>)'
    r'[\s\S]*?'
    r'(</div></div>)',
    lambda m: m.group(1) + "\n" + sales_options_html + "\n        " + m.group(2),
    html
)

# ── 4b. อัพเดท year options (เฉพาะ id="ms-year-options") ────────────────
years = sorted(set(
    str(o["date"].split("/")[2]) for o in orders if len(o["date"]) == 10
))
year_options_html = "\n".join(
    f'          <label class="ms-option"><span>{y}</span>'
    f'<input type="checkbox" value="{y}" onchange="onMsChange(\'ms-year\')" checked></label>'
    for y in years
)
html = re.sub(
    r'(<div class="ms-options"[^>]*id="ms-year-options"[^>]*>)'
    r'[\s\S]*?'
    r'(</div></div>)',
    lambda m: m.group(1) + "\n" + year_options_html + "\n        " + m.group(2),
    html
)

# ── 5. บันทึก ────────────────────────────────────────────────────────────
html = re.sub(r'<!-- source:.*?-->', '', html)
html = html.replace('</head>', f'<!-- source:{xl.name} -->\n</head>', 1)
HTML_PATH.write_text(html, encoding="utf-8")
print(f"Done! Updated: {HTML_PATH.name}")
print(f"Source file:   {xl.name}")
